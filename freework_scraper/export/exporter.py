import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from freework_scraper.models import FreeWorkJob

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions grouped by category
# ---------------------------------------------------------------------------
COLUMNS = [
    # (internal_key, display_label, category)
    ("title", "Titre", "identity"),
    ("company_name", "Entreprise", "identity"),
    ("company_location", "Ville Entreprise", "identity"),
    ("job_category", "Categorie", "identity"),
    ("location", "Localisation", "details"),
    ("remote", "Teletravail", "details"),
    ("salary", "Salaire / TJM", "contract"),
    ("duration", "Duree", "contract"),
    ("experience", "Experience", "contract"),
    ("start_date", "Date de debut", "contract"),
    ("publish_date", "Date de publication", "dates"),
    ("skills", "Competences", "skills"),
    ("sector", "Secteur", "skills"),
    ("description", "Description", "content"),
    ("job_url", "Lien", "meta"),
    ("page_number", "Page", "meta"),
    ("scraped_at", "Scrape le", "meta"),
    ("status", "Statut", "meta"),
]

COLUMN_KEYS = [c[0] for c in COLUMNS]
COLUMN_LABELS = {c[0]: c[1] for c in COLUMNS}
COLUMN_CATEGORIES = {c[0]: c[2] for c in COLUMNS}

# ---------------------------------------------------------------------------
# Excel color scheme — SoClose brand colors
# Primary: #575ECF (purple), Dark: #1b1b1b, Light: #7B80E0
# ---------------------------------------------------------------------------
_HEADER_FILLS = {
    "identity": PatternFill(start_color="575ECF", end_color="575ECF", fill_type="solid"),
    "details": PatternFill(start_color="7B80E0", end_color="7B80E0", fill_type="solid"),
    "contract": PatternFill(start_color="1B1B1B", end_color="1B1B1B", fill_type="solid"),
    "dates": PatternFill(start_color="3D3D3D", end_color="3D3D3D", fill_type="solid"),
    "skills": PatternFill(start_color="4A7FC1", end_color="4A7FC1", fill_type="solid"),
    "content": PatternFill(start_color="86888A", end_color="86888A", fill_type="solid"),
    "meta": PatternFill(start_color="A0A0A0", end_color="A0A0A0", fill_type="solid"),
}

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
FONT_DEFAULT = Font(size=10, name="Calibri")
FONT_LINK = Font(color="575ECF", underline="single", size=10, name="Calibri")
FONT_BOLD = Font(bold=True, size=10, name="Calibri")
FONT_TITLE = Font(bold=True, size=11, name="Calibri", color="1B1B1B")
FONT_DIM = Font(size=9, name="Calibri", color="666666")

FILL_HAS_SALARY = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_NO_SALARY = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
FILL_HAS_REMOTE = PatternFill(start_color="E8E9F7", end_color="E8E9F7", fill_type="solid")
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)

ALIGNMENT_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")

# Recommended column widths
_COL_WIDTHS = {
    "title": 38,
    "company_name": 26,
    "company_location": 22,
    "job_category": 20,
    "location": 22,
    "remote": 18,
    "salary": 24,
    "duration": 16,
    "experience": 18,
    "start_date": 18,
    "publish_date": 18,
    "skills": 40,
    "sector": 20,
    "description": 60,
    "job_url": 45,
    "page_number": 10,
    "scraped_at": 22,
    "status": 10,
}

# ======================================================================
# Public API
# ======================================================================

def export_jobs(
    jobs: list[FreeWorkJob],
    output_dir: str = "output",
    fmt: str = "both",
    search_url: str = "",
) -> list[Path]:
    """Export jobs to CSV and/or Excel. Returns paths of created files."""
    if not jobs:
        logger.warning("No jobs to export.")
        return []

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = _safe_filename(f"freework_jobs_{timestamp}")

    df = _prepare_dataframe(jobs)
    created: list[Path] = []

    if fmt in ("csv", "both"):
        csv_path = out / f"{base}.csv"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        created.append(csv_path)
        logger.info("CSV exported: %s (%d rows)", csv_path, len(df))

    if fmt in ("excel", "both"):
        xlsx_path = out / f"{base}.xlsx"
        _export_excel(df, xlsx_path, search_url)
        created.append(xlsx_path)
        logger.info("Excel exported: %s (%d rows)", xlsx_path, len(df))

    return created

# ======================================================================
# DataFrame preparation
# ======================================================================

def _prepare_dataframe(jobs: list[FreeWorkJob]) -> pd.DataFrame:
    """Flatten jobs into a tabular DataFrame."""
    rows = []
    for j in jobs:
        row = {}
        for key in COLUMN_KEYS:
            val = getattr(j, key, "")
            if val is None:
                row[key] = ""
            elif isinstance(val, int):
                row[key] = val
            else:
                row[key] = str(val)
        rows.append(row)

    df = pd.DataFrame(rows, columns=COLUMN_KEYS)
    df = df.fillna("")

    # Rename columns to display labels
    df.rename(columns=COLUMN_LABELS, inplace=True)

    # Sort: jobs with salary first, then by title
    salary_col = COLUMN_LABELS["salary"]
    title_col = COLUMN_LABELS["title"]
    df["_has_salary"] = df[salary_col].apply(lambda x: 0 if x and str(x).strip() else 1)
    df.sort_values(["_has_salary", title_col], inplace=True)
    df.drop(columns=["_has_salary"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df

# ======================================================================
# Excel export with formatting
# ======================================================================

def _export_excel(df: pd.DataFrame, path: Path, search_url: str = "") -> None:
    """Write a professionally formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FreeWork Jobs"

    num_rows = len(df) + 1  # +1 for header
    num_cols = len(df.columns)

    # --- Row height ---
    ws.row_dimensions[1].height = 30

    # --- Header formatting ---
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        key = COLUMN_KEYS[col_idx - 1]
        category = COLUMN_CATEGORIES.get(key, "meta")

        cell.font = FONT_HEADER
        cell.fill = _HEADER_FILLS.get(category, _HEADER_FILLS["meta"])
        cell.alignment = ALIGNMENT_CENTER
        cell.border = THIN_BORDER

    # --- Column widths ---
    for col_idx in range(1, num_cols + 1):
        key = COLUMN_KEYS[col_idx - 1]
        width = _COL_WIDTHS.get(key, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Precompute column indices ---
    title_col_idx = COLUMN_KEYS.index("title") + 1
    salary_col_idx = COLUMN_KEYS.index("salary") + 1
    remote_col_idx = COLUMN_KEYS.index("remote") + 1
    url_col_idx = COLUMN_KEYS.index("job_url") + 1
    status_col_idx = COLUMN_KEYS.index("status") + 1

    # --- Data rows ---
    for row_idx in range(2, num_rows + 1):
        is_alt_row = (row_idx - 2) % 2 == 1

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = FONT_DEFAULT
            cell.border = THIN_BORDER
            cell.alignment = ALIGNMENT_WRAP

            # Alternating row background
            if is_alt_row:
                cell.fill = FILL_ALT_ROW

        # --- Title column: bold ---
        title_cell = ws.cell(row=row_idx, column=title_col_idx)
        title_cell.font = FONT_TITLE

        # --- Salary cell color coding ---
        salary_cell = ws.cell(row=row_idx, column=salary_col_idx)
        salary_val = str(salary_cell.value or "").strip()
        if salary_val and salary_val != "None":
            salary_cell.fill = FILL_HAS_SALARY
            salary_cell.font = FONT_BOLD
        else:
            salary_cell.fill = FILL_NO_SALARY
            salary_cell.value = ""

        # --- Remote cell color coding ---
        remote_cell = ws.cell(row=row_idx, column=remote_col_idx)
        remote_val = str(remote_cell.value or "").strip()
        if remote_val and remote_val != "None":
            remote_cell.fill = FILL_HAS_REMOTE
        else:
            remote_cell.value = ""

        # --- Clickable job URL ---
        url_cell = ws.cell(row=row_idx, column=url_col_idx)
        url_val = str(url_cell.value or "").strip()
        if url_val.startswith("http"):
            url_cell.hyperlink = url_val
            url_cell.font = FONT_LINK

        # --- Status column: color coding ---
        status_cell = ws.cell(row=row_idx, column=status_col_idx)
        status_val = str(status_cell.value or "").strip().lower()
        if status_val == "ok":
            status_cell.fill = FILL_OK
        elif status_val == "error":
            status_cell.fill = FILL_ERROR
        status_cell.alignment = ALIGNMENT_CENTER
        status_cell.font = FONT_DIM

        # --- Clean remaining "None" values ---
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None or str(cell.value).strip() == "None":
                cell.value = ""

    # --- Freeze panes (header row + first column) ---
    ws.freeze_panes = "B2"

    # --- Auto filter ---
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)

    logger.info("Excel file written: %s (%d jobs)", path, len(df))

# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------
def _add_summary_sheet(writer, df: pd.DataFrame, search_url: str) -> None:
    """Add a summary/statistics sheet to the Excel file."""
    wb = writer.book
    ws = wb.create_sheet("Resume", 0)

    # Title
    ws.merge_cells("A1:D1")
